r"""
15.03.2017 - версия 0.01
Диалог для запуска расчетов объемов работ. Для работы в текущей папке должен лежать файл "ВР шаблон.docx" 
Компиляция командой: pyinstaller --path="C:\Users\milkov\AppData\Local\Application Data\Programs\Python\Python35\Lib\site-packages\PyQt5\Qt\bin"\ -F gen5.py 
16.03.2017 - версия 0.02
Переименовал проект в gen5. Разложил по папкам. Добавил эмблему лигато на форму.
17.03.2017 - версия 0.03
Добавил иконку. Добавил формирование заказных спецификаций
20.03.2017 - версия 0.04
Добавил смену курсора при длительной обработке. Добавил формирование обычных cпецификаций в автокаде
22.03.2017 - версия 0.05
Работа с Excel через класс
24.03.2017 - версия 0.06
Справка по работе
27.04.2017 - версия 0.07
Работа с эксель через openpyxl
21.06.2017 - версия 0.08
Добавлен прогресс бар выполнения задачи
23.06.2017 - версия 0.09
Добавлен переход на сайт лигато.рф при нажатии на картинку
"""

import sys
import os
import webbrowser

from PyQt5.QtWidgets import (QWidget, QMainWindow, QLabel, QPushButton, QLineEdit, QFileDialog, QMessageBox, QTextEdit, QGridLayout, QApplication, QListWidget, QProgressBar)
from PyQt5.QtCore import QDir, Qt
from PyQt5.QtGui import QPixmap, QIcon
import openpyxl

from mdvlib.mso import Excel

from vr import open_and_calc
from so import make_so
from so_acad import make_spec

VERSION = "0.09"

help_text = """Данная программа предназначена для расчета Объемов работ, Спецификаций и Заказных Спецификаций.
Входными данными для расчета является Excel файл, пример которого VR_пример.xlsx можно найти в папке templates. Также в папке templates лежат
шаблоны для вывода готовых расчетов. Готовые расчеты выводятся в папку Готовое.

Интерфейс.
Загрузить проект - перед началом работы надо нажать на эту кнопку и выбрать Excel файл с данными для расчетов.
Объемы работ - создает ведомость объемов работ. Готовый файл называется 'ВР готовый.docx' и лежит в папке Готовое.
Справка - показать данную справку.
Выход - выход  из программы.
СО1 - создает Спецификация оборудования, изделий и материалов поставки ООО \"Газпром комплектация\". Готовый файл называется 'СО1 готовый.docx' и лежит в папке Готовое.
СО2 - создает Спецификация оборудования, изделий и материалов поставки АО \"Газпром СтройТЭК Салават\". Готовый файл называется 'СО2 готовый.docx' и лежит в папке Готовое.
СО3 - создает Спецификация оборудования, изделий и материалов поставки подрядчика. Готовый файл называется 'СО3 готовый.docx' и лежит в папке Готовое.
Спецификация - создает обычную спецификацию в формает AutoCAD на листах формата А3 для выбранного из списка элемента. Готовый файл называется 'С готовый.dwg' и лежит в папке Готовое.
Область выбора компонентов - выбор компонента для которого будет создаваться Спецификация.

Формат входного xlsx файла.
Файл должен содержать 3 листа: ЛИ, ИГ, СП.
Формат данных на листе СП: 
первые 2 строки - не используются, 3 строка с 1 по 20 колонку - названия параметров, с 21 и далее компоненты спецификаций
кол.1 -  Поз. - позиция на чертеже
кол.2 -  Обозначение - обозначение
кол.3 -  Наименование - наименование
кол.4 -  Кол. - количество, для всего что учавстсвует в линиях(см. ниже), для труб - длина в м, считается автоматически
кол.5 -  Масса ед - масса в кг
кол.6 -  Примечание - примечание (единицы измерения указывать не нужно)
кол.7 -  Завод изготовитель - завод
кол.8 -  Единица измерения - шт., м, кг, т и тд
кол.9 -  Type - 1 символ кодирования типа детали (смотри Коды типов деталей)
кол.10 - TAG - уникальный идентификатор детали. см 'Правила именования TAG'
кол.11 - Dy - условный диаметр, если есть
кол.12 - D - диаметр детали, мм, если есть
кол.13 - s - толщина стенки детали, мм, если есть
кол.14 - Dy2 - условный диаметр отвлетвления для тройника и перехода для перехода
кол.15 - D2 - диаметр отвлетвления для тройника и перехода для перехода в мм
кол.16 - Длина - длина детали в мм, для труб указать 1000
кол.17 - SubType - для отводов угол поворота в градусах
кол.18 - СО - номер спецификации к которой относится данная деталь (1-3)
кол.19 - Площадь - площадь поверхности детали, м2, если есть
кол.20 - Масса с изоляцией - масса детали в заводской изоляции
кол.21 и далее - Название компонента. Например у нас проект состоит из основного чертежа, чертежа свечи и чертежа стояка. Назовем Колонку 21 как Основная, 22 - Свеча,
23 - Стояк. В колонки 22 и 23 вручную внесем количество деталей. в колонку 21 запишем формулу кол4-кол22-кол23.
строки 4 и далее используются для записей деталей, по 1 строке на каждую деталь, для удобства чтения можно вертикально объединять ячейки.

Для описания объекта используем деление его на линии, которые представляют собой связанные друг с другом последовательно детали трубопровода.

Формат данных на листе ЛИ:
На данном листе описываем все линии трубопровода(кроме линий относящихся к импульсному газу). 
Описание каждой новой линии начинается с шапки вида:
кол.1 - линия - ключевое слово означающее что начинается описание новой линии трубопровода, 
кол.2 - L-1-1 - название линии (не должно повторяться), 
кол.3 - 7.4 - рабочее давление МПа в линии, 
кол 4 - код учета стыковки данной линии с чем либо: 0 - считаем оба стыка(начало линии и конец линии), 1 - считаем только последний стык, 2 - считаем только первый стык, 3 - не считаем стыки
кол.5 - TAG того с чем стыкуеся начало линии (или RAM - если это гарантийный стык), если ни с чем не стыкуется оствить пустым
кол.6 - TAG того с чем стыкуеся конец линии (или RAM - если это гарантийный стык), если ни с чем не стыкуется оствить пустым
Далее идет описание деталей из которых состоит линия в формате:
кол.1 - TAG детали (см. лист СП)
кол.2 - p - если деталь подземная и n - если деталь надземная
кол.3 - для труб - длина в мм, для остального оставить пустым

Формат данных на листе ИГ:
На данном листе описываем все линии относящиеся к импульсному газу. 
Формат точно такой-же как у листа ЛИ.

Коды типов деталей (колонка Type)
k - краны
t - трубы 
o - отводы, 
r - тройники, 
p - переходы, 
d - днища,  
z - заглушки, 
f - фланцы, 
a - фланцевые заглушки
l - все остальное что нужно писать в ВР как установка/монтаж, 
для болтов, гаек, прокладок, шпилек, пробок, красок, бетона и тп код типа не пишем

Правила именования поля TAG 
первый символ идентификатора должен совпадать с Type. Исключения: nsv14 и shtuz - Ниппельное соединение ввертное, Штуцер приварной эти идентификаторы нельзя пока что менять

"""
    
class HelpWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Справка') 
        self.textEdit = QTextEdit()
        self.textEdit.setReadOnly(True)
        self.setCentralWidget(self.textEdit)
        self.setGeometry(50, 50, 1000, 500)
        self.textEdit.setText(help_text)        
  

class Example(QWidget):
    
    def __init__(self):
        super().__init__()
        
        self.initUI()
        
    def initUI(self):
        
        self.curdir = os.path.dirname(os.path.abspath(__file__))
        
        self.setWindowIcon(QIcon(os.path.join(self.curdir, "Pictures", "icon.ico")))
        
        title = QLabel('Проект:')
        self.titleEdit = QLabel('Проект не загружен')
        
        self.ligato = QLabel('')
        myPix = QPixmap(os.path.join(self.curdir, "Pictures", "ligato.png"))
        self.ligato.setPixmap(myPix)
        self.ligato.mousePressEvent = self.ligato_site

        self.btn_load = QPushButton('Загрузить проект', self)
        self.btn_load.clicked.connect(self.showDlgLoadExl)

        self.btn_vr = QPushButton('Объемы работ', self)
        self.btn_vr.setDisabled(True)
        self.btn_vr.clicked.connect(self.calcvr)

        self.btn_so1 = QPushButton('CO1', self)
        self.btn_so1.setDisabled(True)
        self.btn_so1.clicked.connect(self.makeso)

        self.btn_so2 = QPushButton('CO2', self)
        self.btn_so2.setDisabled(True)
        self.btn_so2.clicked.connect(self.makeso)

        self.btn_so3 = QPushButton('CO3', self)
        self.btn_so3.setDisabled(True)
        self.btn_so3.clicked.connect(self.makeso)

        self.btn_spec = QPushButton('Спецификация', self)
        self.btn_spec.setDisabled(True)
        self.btn_spec.clicked.connect(self.makes)
        
        self.list_det = QListWidget(self)

        self.btn_exit = QPushButton('Выход', self)
        self.btn_exit.clicked.connect(QApplication.closeAllWindows)

        self.btn_help = QPushButton('Справка', self)
        self.btn_help.clicked.connect(self.help)

        self.progress = QProgressBar(self)
        self.progress.setValue(0)

        grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(title, 1, 0)
        grid.addWidget(self.titleEdit, 1, 1, 1, 3)
        grid.addWidget(self.btn_load, 2, 0)
        grid.addWidget(self.list_det, 2, 2, 2, 2)
        grid.addWidget(self.btn_spec, 5, 1)
        grid.addWidget(self.btn_vr, 3, 0)
        grid.addWidget(self.btn_so1, 2, 1)
        grid.addWidget(self.btn_so2, 3, 1)
        grid.addWidget(self.btn_so3, 4, 1)
        grid.addWidget(self.btn_exit, 5, 0)
        grid.addWidget(self.btn_help, 4, 0)
        grid.addWidget(self.ligato, 4, 3, 2, 1)
        grid.addWidget(self.progress, 6, 0, 1, 4)
        
        self.setLayout(grid) 
        
        self.setWindowTitle('Генератор объемов работ и спецификаций. версия %s' % VERSION)    

        self.hw = HelpWindow()

        self.show()

    def ligato_site(self, event):
        webbrowser.open('http://xn--80afolsx.xn--p1ai/')

    def showDlgLoadExl(self):
        fname = QFileDialog().getOpenFileName(self, 'Open file', self.curdir, "Excel files (*.xlsx)")

        if fname[0]:
            self.titleEdit.setText(fname[0])
            self.btn_vr.setDisabled(False)
            self.btn_so1.setDisabled(False)
            self.btn_so2.setDisabled(False)
            self.btn_so3.setDisabled(False)
            self.btn_spec.setDisabled(False)
            self.list_det.clear()
            wb = openpyxl.load_workbook(fname[0])
            ws = wb['СП']
            for col in range(21, ws.max_column):
                if ws.cell(row=3, column=col).value=="":
                    break
                else:
                    self.list_det.addItem(ws.cell(row=3, column=col).value)
            self.list_det.setCurrentRow(0)
            wb.close()
        
    def calcvr(self):
        QApplication.setOverrideCursor(Qt.WaitCursor)
        open_and_calc(self.titleEdit.text(), self.curdir, self.progress)
        QApplication.restoreOverrideCursor()
        msgBox = QMessageBox(QMessageBox.Information, "Внимание", "Объемы работ были успешно созданы.")
        msgBox.exec_()
        self.progress.setValue(0)

    def makeso(self):
        QApplication.setOverrideCursor(Qt.WaitCursor)
        sender = self.sender()
        make_so(self.titleEdit.text(), self.curdir, sender.text()[2], self.progress)
        QApplication.restoreOverrideCursor()
        msgBox = QMessageBox(QMessageBox.Information, "Внимание", "Спецификация оборудования СО%s успешно создана." % sender.text()[2])
        msgBox.exec_()
        self.progress.setValue(0)
    
    def makes(self):
        QApplication.setOverrideCursor(Qt.WaitCursor)
        make_spec(self.titleEdit.text(), self.curdir, 21+self.list_det.currentRow(), self.progress)
        QApplication.restoreOverrideCursor()
        msgBox = QMessageBox(QMessageBox.Information, "Внимание", "Спецификация для %s была успешно создана." % self.list_det.currentItem().text())
        msgBox.exec_()
        self.progress.setValue(0)

    def help(self):
        self.hw.show()

    
def main():
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())
    
if __name__ == '__main__':
    main()    