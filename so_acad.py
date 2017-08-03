"""
20.03.2017 - версия 0.01
Считывает данные из xls спецификации для выбранной детали и выводит
спецификацию в фомате автокад.
Добавлена вставка шаблона А3 для каждого листа спецификации, вывод штук ЦЕЛЫМИ,
Вывод ед. измерения. в графе примечания.
21.03.2017 - версия 0.02
Доделал точный перенос на след страницу
22.03.2017 - версия 0.03
Работа с Excel через класс
20.06.2017 - версия 0.04
Работа с эксель через openpyxl
21.06.2017 - версия 0.05
Добавлен прогресс бар выполнения задачи
"""

import os
from pyautocad import Autocad, APoint, ACAD
import openpyxl

from mdvlib.util import split_str_space


def make_spec(fname, file_path, col_num, progress):
    POZ_COL = 1
    TIP_COL = 2
    NAME_COL = 3
    MASSA_COL = 5
    PRIM_COL = 6
    ED_COL = 8
    NUM_COL = int(col_num)
    TIP_LIM = 32
    NAME_LIM = 35
    PRIM_LIM = 12

    acad = Autocad(create_if_not_exists=True)
    adoc = acad.app.Documents.Open(os.path.join(
        file_path, 'templates', "Шаблон.dwg"))
    adoc.SaveAs(os.path.join(file_path, 'Готовое',
                             "C готовый.dwg"), ACAD.ac2007_dwg)

    def put_s_acad_line(l, x, y):
        for i in range(6):
            if l[i]:
                p = APoint(x + dx[i], y, 0)
                text = acad.model.AddText(l[i], p, 2.5)
                if i in (0, 3, 4, 5):
                    text.Alignment = ACAD.acAlignmentCenter
                    text.TextAlignmentPoint = p

    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb['СП']

    spec_lines = []
    # от 0 до 100 читаем данные (это первые 50%)
    progress.setMaximum(200)
    for i in range(4, ws.max_row):
        progress.setValue(100 * i / ws.max_row)
        if not ws.cell(row=i, column=NUM_COL).value:
            continue
        l = ["", "", "", "", "", ""]
        if ws.cell(row=i, column=POZ_COL).value:
            l[0] = "%d" % int(ws.cell(row=i, column=POZ_COL).value)
        l_1 = split_str_space(ws.cell(row=i, column=TIP_COL).value, TIP_LIM)
        if l_1:
            l[1] = l_1.pop(0)
        l_2 = split_str_space(ws.cell(row=i, column=NAME_COL).value, NAME_LIM)
        if l_2:
            l[2] = l_2.pop(0)
        ed = ws.cell(row=i, column=ED_COL).value
        if ed == "м":
            l[3] = "%.1f" % float(ws.cell(row=i, column=NUM_COL).value)
        else:
            l[3] = "%d" % int(ws.cell(row=i, column=NUM_COL).value)
        if ws.cell(row=i, column=MASSA_COL).value:
            l[4] = "%.1f" % ws.cell(row=i, column=MASSA_COL).value
        prim = ws.cell(row=i, column=PRIM_COL).value
        if not prim:
            prim = ""
        l_5 = split_str_space(prim + " " + ed, PRIM_LIM)
        if l_5:
            l[5] = l_5.pop(0)

        spec_lines.append(l)

        while l_1 or l_2 or l_5:
            l = ["", "", "", "", "", ""]
            if l_1:
                l[1] = l_1.pop(0)
            if l_2:
                l[2] = l_2.pop(0)
            if l_5:
                l[5] = l_5.pop(0)
            spec_lines.append(l)

    wb.close()

    # Вывод на листы формата А3 автокада
    x0 = 0  # начальная точка вставки
    y0 = 270  # начальная точка вставки
    long = True  # левая сторона (длинная)
    yf = y0  # начальная точка вставки (копия)
    dy = 8  # расстояние между строками
    dx = [27.5, 36, 96, 165, 177.5, 195]  # смещение столбцов относительно x0

    long_tab = 248  # высота левой(длинной) таблицы
    short_tab = 192  # высота правой(короткой) таблицы

    l_len = len(spec_lines) - 1
    for i, l in enumerate(spec_lines):
        # от 100 до 200 выводим данные (это вторые 50%)
        progress.setValue(100 + 100 * i / l_len)
        put_s_acad_line(l, x0, y0)
        y0 = y0 - dy
        if long and yf - long_tab > y0:
            y0 = yf
            x0 = x0 + 210
            long = False
        elif not long and yf - short_tab > y0:
            y0 = yf
            x0 = x0 + 220
            long = True
            p = APoint(x0, 0)
            acad.model.InsertBlock(p, 'A3_forma_s', 1, 1, 1, 0)
    adoc.Close(True)


if __name__ == "__main__":
    pass
