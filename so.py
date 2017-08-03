"""
22.03.2017 - версия 0.01
Вывод заказных спецификаций в фомате Word
20.06.2017 - версия 0.02
Работа с эксель через openpyxl
21.06.2017 - версия 0.03
Добавлен прогресс бар выполнения задачи
"""

import sys
import os
import win32com.client
import openpyxl

from mdvlib.mso import CH_2, CH_3, CH_D, CH_GRAD
from mdvlib.util import split_str_space, float_or_none, int_or_none

def make_so(fname, file_path, so_num, progress):
    TAG_COL = 10; SO_COL = 18; NAME_COL = 3; TIP_COL = 2; POST_COL = 7; ED_COL = 8; NUM_COL = 4; MASSA_COL = 5; IZ_MAS_COL = 20
    NAME_LIM = 66; TIP_LIM = 32; POST_LIM = 23

    #делаем доступными вордовские константы
    win32com.client.gencache.EnsureDispatch('Word.Application')
    wordapp = win32com.client.Dispatch("Word.Application") # Create new Word Object
    worddoc = wordapp.Documents.Open(os.path.join(file_path, 'templates', "СО шаблон.docx")) # Create new Document Object
    worddoc.SaveAs2(os.path.join(file_path, 'Готовое', "CO%s готовый.docx" % so_num))

    tab = worddoc.Tables.Item(1)
    
    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb['СП']

    row = 2
    so_num=int(so_num)
    progress.setMaximum(ws.max_row-1)
    for i in range(4, ws.max_row):
        progress.setValue(i)
        so = ws.cell(row=i, column=SO_COL).value
        if so==so_num:
            name = ws.cell(row=i, column=NAME_COL).value 
            tip = ws.cell(row=i, column=TIP_COL).value 
            post = ws.cell(row=i, column=POST_COL).value 
            ed = []
            ed.append(ws.cell(row=i, column=ED_COL).value)
            num = []
            if ed[0]=="м":
                num.append("%.2f" % float_or_none(ws.cell(row=i, column=NUM_COL).value))
            else:
                num.append("%d" % int_or_none(ws.cell(row=i, column=NUM_COL).value))
            iz_mas = float_or_none(ws.cell(row=i, column=IZ_MAS_COL).value)
            massa = []
            m = float_or_none(ws.cell(row=i, column=MASSA_COL).value)
            if m:
                massa.append("%.1f" % m)
            
            #для труб посчитать тонны
            if ed[0]=="м":
                ed.append("т")
                num.append("%.1f" % (float(num[0]) * float(massa[0]) / 1000) )
            #если есть масса в изоляции - вывести ее
            if iz_mas:
               if so_num == 1:
                    #для трубы выноска 3) для остального выноска 4)
                    if ed[0]=="м":
                        vin = " 3_)"
                    else:
                        vin = " 4_)"
                    massa[0] = massa[0] + vin
                    massa.append ("%.1f 5_)" % (iz_mas))
               elif so_num == 3:
                    # выноска 3) и выноска 4)
                    massa[0] = massa[0] + " 3_)"
                    massa.append ("%.1f 4_)" % (iz_mas))
            
            max_idx = 0

            idx = 0
            for j in split_str_space(name, NAME_LIM):
                tab.Cell(row + idx, 2).Range.Text = j
                idx = idx + 1
            if idx > max_idx:
                max_idx = idx

            idx = 0
            for j in split_str_space(tip, TIP_LIM):
                tab.Cell(row + idx, 3).Range.Text = j
                idx = idx + 1
            if idx > max_idx:
                max_idx = idx

            idx = 0
            for j in split_str_space(post, POST_LIM):
                tab.Cell(row + idx, 5).Range.Text = j
                idx = idx + 1
            if idx > max_idx:
                max_idx = idx

            idx = 0
            for j in ed:
                tab.Cell(row + idx, 6).Range.Text = j
                idx = idx + 1
            if idx > max_idx:
                max_idx = idx

            idx = 0
            for j in num:
                tab.Cell(row + idx, 7).Range.Text = j
                idx = idx + 1
            if idx > max_idx:
                max_idx = idx
            idx = 0

            for j in massa:
                tab.Cell(row + idx, 8).Range.Text = j
                idx = idx + 1
            if idx > max_idx:
                max_idx = idx

            row = row + max_idx + 1

    wb.close()

    myrange = worddoc.Content
    while myrange.Find.Execute(FindText="3_)"):
        myrange.Font.Superscript = True
        myrange.Text = "3)"
        myrange = worddoc.Content
    
    myrange = worddoc.Content
    while myrange.Find.Execute(FindText="4_)"):
        myrange.Font.Superscript = True
        myrange.Text = "4)"
        myrange = worddoc.Content
    
    myrange = worddoc.Content
    while myrange.Find.Execute(FindText="5_)"):
        myrange.Font.Superscript = True
        myrange.Text = "5)"
        myrange = worddoc.Content
            
    worddoc.Save()
    worddoc.Close()

if __name__ == "__main__":
    pass
